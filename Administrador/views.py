"""
Vistas del panel administrativo de Agrophia.

Este archivo centraliza operaciones de administración para:
- Tiendas
- Productos
- Usuarios
- Pedidos
- Reportes (Excel e impresión)
- Actividad reciente

Nota:
El archivo conserva una organización histórica por bloques funcionales.
Los comentarios agregados delimitan secciones y explican flujos críticos.
"""

# ============================================================
# BLOQUE 1: GESTION DE TIENDAS (ADMIN)
# ============================================================
from Tiendas.models import Shop
from django.views.decorators.http import require_POST

def store_admin_view(request):
    """Renderiza la tabla administrativa de tiendas.

    Args:
        request: Solicitud HTTP del panel admin.

    Returns:
        HttpResponse con la plantilla de listado de tiendas.
    """
    # Lista general de tiendas ordenadas por creación reciente.
    tiendas = Shop.objects.all().order_by('-created_at')
    return render(request, 'administrador/store_admin.html', {'tiendas': tiendas})

def tienda_admin_detalle_view(request, tienda_id):
    """Muestra el detalle de una tienda y de su propietario asociado.

    Args:
        request: Solicitud HTTP.
        tienda_id: ID de la tienda a consultar.
    """
    # Detalle completo de tienda + datos del propietario (si existe).
    tienda = get_object_or_404(Shop, id=tienda_id)
    usuario = tienda.owner if tienda.owner else None
    usuario_info = None
    if usuario:
        usuario_info = Register.objects.filter(id_usuario=usuario.id).first()
    return render(request, 'administrador/tienda_detalle_admin.html', {
        'tienda': tienda,
        'usuario': usuario,
        'usuario_info': usuario_info,
    })

def tienda_admin_crear_view(request):
    """Crea una tienda desde administración enlazándola a un cliente existente.

    También prepara datos para autocompletado del formulario.
    """
    # Crea una tienda enlazándola a un usuario cliente disponible.
    from usuarios.models import Register
    success = False
    errores = {}
    valores = {}
    usuarios_con_tienda = Shop.objects.exclude(owner__isnull=True).values_list('owner__id', flat=True)
    usuarios_disponibles = Register.objects.exclude(id_usuario__in=usuarios_con_tienda).exclude(estado='admin').order_by('nombres', 'apellidos')
    usuarios_disponibles_data = [
        {
            'id_usuario': usuario.id_usuario,
            'nombres': usuario.nombres,
            'apellidos': usuario.apellidos,
            'correo_electronico': usuario.correo_electronico,
            'telefono': usuario.telefono,
            'departamento': usuario.departamento,
            'municipio': usuario.municipio,
            'direccion_completa': usuario.direccion_completa,
        }
        for usuario in usuarios_disponibles
    ]

    if request.method == 'POST':
        owner_id = request.POST.get('owner_id', '').strip()
        nombre = request.POST.get('nombre', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        email = request.POST.get('email', '').strip()
        departamento = request.POST.get('departamento', '').strip()
        municipio = request.POST.get('municipio', '').strip()
        direccion = request.POST.get('direccion', '').strip()
        horario = request.POST.get('horario', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        valores = {
            'owner_id': owner_id,
            'nombre': nombre,
            'telefono': telefono,
            'email': email,
            'departamento': departamento,
            'municipio': municipio,
            'direccion': direccion,
            'horario': horario,
            'descripcion': descripcion,
        }
        # Validaciones
        propietario = usuarios_disponibles.filter(id_usuario=owner_id).first() if owner_id else None
        if not owner_id:
            errores['owner_id'] = 'Debes seleccionar un usuario cliente existente.'
        elif not propietario:
            errores['owner_id'] = 'El usuario seleccionado no esta disponible para crear tienda.'
        if not nombre:
            errores['nombre'] = 'El nombre es obligatorio.'
        if not telefono:
            errores['telefono'] = 'El teléfono es obligatorio.'
        if not email:
            errores['email'] = 'El correo es obligatorio.'
        if not departamento:
            errores['departamento'] = 'El departamento es obligatorio.'
        if not municipio:
            errores['municipio'] = 'El municipio es obligatorio.'
        # Validar que el municipio pertenezca al departamento
        from .departamentos_municipios import DEPARTAMENTOS_MUNICIPIOS
        if departamento and municipio:
            municipios_validos = DEPARTAMENTOS_MUNICIPIOS.get(departamento)
            if municipios_validos and municipio not in municipios_validos:
                errores['municipio'] = f'El municipio "{municipio}" no corresponde al departamento seleccionado.'
        if not errores:
            Shop.objects.create(
                owner_id=propietario.id_usuario,
                nombre=nombre,
                telefono=telefono,
                email=email,
                departamento=departamento,
                municipio=municipio,
                direccion=direccion,
                horario=horario,
                descripcion=descripcion,
                is_active=True
            )
            success = True
            valores = {}
    from .departamentos_municipios import DEPARTAMENTOS_MUNICIPIOS
    return render(request, 'administrador/tienda_crear_admin.html', {
        'success': success,
        'errores': errores,
        'valores': valores,
        'usuarios_disponibles': usuarios_disponibles,
        'usuarios_disponibles_data': usuarios_disponibles_data,
        'departamentos_municipios': DEPARTAMENTOS_MUNICIPIOS,
    })

def tienda_admin_editar_view(request, tienda_id):
    """Edita información de una tienda existente.

    Incluye validación de municipio contra departamento.
    """
    # Edita datos administrativos de tienda con validación básica de ubicación.
    tienda = get_object_or_404(Shop, id=tienda_id)
    success = False
    errores = {}
    valores = {
        'nombre': tienda.nombre,
        'telefono': tienda.telefono,
        'email': tienda.email,
        'departamento': tienda.departamento,
        'municipio': tienda.municipio,
        'direccion': tienda.direccion,
        'horario': tienda.horario,
        'descripcion': tienda.descripcion,
    }
    from .departamentos_municipios import DEPARTAMENTOS_MUNICIPIOS
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        email = request.POST.get('email', '').strip()
        departamento = request.POST.get('departamento', '').strip()
        municipio = request.POST.get('municipio', '').strip()
        direccion = request.POST.get('direccion', '').strip()
        horario = request.POST.get('horario', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        valores = {
            'nombre': nombre,
            'telefono': telefono,
            'email': email,
            'departamento': departamento,
            'municipio': municipio,
            'direccion': direccion,
            'horario': horario,
            'descripcion': descripcion,
        }
        # Validaciones
        if not nombre:
            errores['nombre'] = 'El nombre es obligatorio.'
        if not telefono:
            errores['telefono'] = 'El teléfono es obligatorio.'
        if not email:
            errores['email'] = 'El correo es obligatorio.'
        if not departamento:
            errores['departamento'] = 'El departamento es obligatorio.'
        if not municipio:
            errores['municipio'] = 'El municipio es obligatorio.'
        # Validar que el municipio pertenezca al departamento
        from .departamentos_municipios import DEPARTAMENTOS_MUNICIPIOS
        if departamento and municipio:
            municipios_validos = DEPARTAMENTOS_MUNICIPIOS.get(departamento)
            if municipios_validos and municipio not in municipios_validos:
                errores['municipio'] = f'El municipio "{municipio}" no corresponde al departamento seleccionado.'
        if not errores:
            tienda.nombre = nombre
            tienda.telefono = telefono
            tienda.email = email
            tienda.departamento = departamento
            tienda.municipio = municipio
            tienda.direccion = direccion
            tienda.horario = horario
            tienda.descripcion = descripcion
            tienda.save()
            success = True
        return render(request, 'administrador/tienda_editar_admin.html', {
            'success': success,
            'errores': errores,
            'valores': valores,
            'departamentos_municipios': DEPARTAMENTOS_MUNICIPIOS,
            'tienda': tienda,
        })
    return render(request, 'administrador/tienda_editar_admin.html', {
        'success': success,
        'errores': errores,
        'valores': valores,
        'departamentos_municipios': DEPARTAMENTOS_MUNICIPIOS,
        'tienda': tienda,
    })

@require_POST
def tienda_admin_block_view(request, tienda_id):
    """Bloquea (desactiva) una tienda por ID."""
    # Deshabilita tienda (soft-disable) para ocultarla de flujos activos.
    tienda = get_object_or_404(Shop, id=tienda_id)
    tienda.is_active = False
    tienda.save(update_fields=["is_active"])
    return JsonResponse({'ok': True})

@require_POST
def tienda_admin_unblock_view(request, tienda_id):
    """Desbloquea (activa) una tienda por ID."""
    # Rehabilita tienda previamente bloqueada.
    tienda = get_object_or_404(Shop, id=tienda_id)
    tienda.is_active = True
    tienda.save(update_fields=["is_active"])
    return JsonResponse({'ok': True})

# ============================================================
# BLOQUE 2: REPORTES DE TIENDAS Y PRODUCTOS
# ============================================================

# Reporte de tiendas
def reporte_tiendas_view(request):
    """Genera reporte de tiendas en formato Excel o imprimible."""
    # Exporta reporte de tiendas en formato Excel o vista imprimible.
    output_format = request.GET.get('output_format', 'excel')
    tiendas = Shop.objects.all().order_by('-created_at')
    headers = ['ID', 'Nombre', 'Teléfono', 'Correo', 'Departamento', 'Municipio', 'Dirección', 'Horario', 'Estado', 'Fecha creación']
    rows = [
        [
            t.id,
            t.nombre,
            t.telefono,
            t.email,
            t.departamento,
            t.municipio,
            t.direccion,
            t.horario,
            'Activo' if t.is_active else 'Inactivo',
            t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else ''
        ]
        for t in tiendas
    ]

    if output_format == 'print':
        return render_generic_report_print(request, 'Reporte de tiendas', headers, rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tiendas Agrophia'
    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='16A34A', end_color='16A34A', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        ws.append(row)
    autofit_worksheet_columns(ws)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="tiendas_agrophia.xlsx"'
    wb.save(response)
    return response
# Reporte de productos
def reporte_productos_view(request):
    """Genera reporte de productos en formato Excel o imprimible."""
    # Exporta reporte de productos con formato equivalente a tiendas.
    from Productos.models import Product
    output_format = request.GET.get('output_format', 'excel')
    productos = Product.objects.all().order_by('-created_at')
    headers = ['ID', 'Nombre', 'Tipo', 'Unidad', 'Precio', 'Descripción', 'Garantía', 'Estado', 'Fecha creación']
    rows = [
        [
            p.id,
            p.nombre,
            f"{p.tipo} ({p.tipo_otro})" if p.tipo == 'Otros' and p.tipo_otro else p.tipo,
            p.unidad,
            str(p.precio),
            p.descripcion,
            p.garantia,
            'Activo' if p.is_active else 'Inactivo',
            p.created_at.strftime('%Y-%m-%d %H:%M') if hasattr(p, 'created_at') and p.created_at else ''
        ]
        for p in productos
    ]

    if output_format == 'print':
        return render_generic_report_print(request, 'Reporte de productos', headers, rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Productos Agrophia'
    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        ws.append(row)
    autofit_worksheet_columns(ws)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="productos_agrophia.xlsx"'
    wb.save(response)
    return response
# ============================================================
# BLOQUE 3: PRODUCTOS (DETALLE / CREACION / EDICION / ESTADO)
# ============================================================

# Vista detalle producto admin
def producto_admin_detalle_view(request, product_id):
    """Muestra el detalle administrativo de un producto."""
    # Muestra ficha detallada del producto para revisión administrativa.
    from Productos.models import Product
    producto = get_object_or_404(Product, id=product_id)
    return render(request, 'administrador/producto_detalle_admin.html', {'producto': producto})
from Productos.models import Product, ProductImage
from django.core.exceptions import ValidationError
from decimal import Decimal, InvalidOperation
# --- Vistas para crear y editar productos desde el admin ---
def producto_admin_crear_view(request):
    """Crea un producto desde el panel admin con validaciones completas.

    Controla:
    - Campos obligatorios.
    - Rango/precio válido.
    - Tipos de imagen permitidos y límite de cantidad.
    """
    # Alta administrativa de producto con validaciones de contenido e imágenes.
    success = False
    errores = {}
    valores = {}
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        tipo = request.POST.get('tipo', '').strip()
        tipo_otro = request.POST.get('tipo_otro', '').strip()
        unidad = request.POST.get('unidad', '').strip()
        precio = request.POST.get('precio', '').strip()
        stock = request.POST.get('stock', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        garantia = request.POST.get('garantia', '').strip()
        fotos = request.FILES.getlist('fotos')
        valores = {
            'nombre': nombre,
            'tipo': tipo,
            'tipo_otro': tipo_otro,
            'unidad': unidad,
            'precio': precio,
            'stock': stock,
            'descripcion': descripcion,
            'garantia': garantia,
        }
        # Validaciones básicas
        if not nombre:
            errores['nombre'] = 'El nombre es obligatorio.'
        if not tipo:
            errores['tipo'] = 'El tipo es obligatorio.'
        if tipo == 'Otros' and not tipo_otro:
            errores['tipo_otro'] = 'Debes especificar el tipo.'
        if not unidad:
            errores['unidad'] = 'La unidad es obligatoria.'
        if not precio:
            errores['precio'] = 'El precio es obligatorio.'
        else:
            try:
                precio_val = Decimal(precio)
                if precio_val <= 0:
                    errores['precio'] = 'El precio debe ser mayor que 0.'
            except (InvalidOperation, ValueError):
                errores['precio'] = 'Precio inválido.'
        if not stock:
            errores['stock'] = 'La cantidad disponible es obligatoria.'
        else:
            try:
                stock_val = int(stock)
                if stock_val < 0:
                    errores['stock'] = 'La cantidad disponible no puede ser negativa.'
            except (TypeError, ValueError):
                errores['stock'] = 'Cantidad disponible inválida.'
        if not descripcion:
            errores['descripcion'] = 'La descripción es obligatoria.'
        if not garantia:
            errores['garantia'] = 'La garantía es obligatoria.'
        if not fotos:
            errores['fotos'] = 'Debes cargar al menos una imagen.'
        elif len(fotos) > 8:
            errores['fotos'] = 'Solo puedes cargar máximo 8 imágenes.'
        else:
            for photo in fotos:
                if not (photo.content_type or '').startswith('image/'):
                    errores['fotos'] = 'Solo se permiten archivos de imagen.'
                    break
        # Validación de modelo
        if not errores:
            producto = Product(
                nombre=nombre,
                tipo=tipo,
                tipo_otro=tipo_otro,
                unidad=unidad,
                precio=precio,
                stock=stock,
                descripcion=descripcion,
                garantia=garantia,
                is_active=(int(stock) > 0),
                owner=request.user if request.user.is_authenticated else None
            )
            try:
                producto.full_clean()
            except ValidationError as exc:
                for field, messages in exc.message_dict.items():
                    errores[field] = messages[0] if messages else 'Valor inválido.'
            else:
                producto.save()
                for photo in fotos:
                    ProductImage.objects.create(product=producto, image=photo)
                success = True
                valores = {}
    return render(request, 'administrador/producto_crear_admin.html', {
        'success': success,
        'errores': errores,
        'valores': valores,
        'tipo_choices': Product.TIPO_CHOICES,
        'unidad_choices': Product.UNIDAD_CHOICES,
    })

def producto_admin_editar_view(request, product_id):
    """Edita un producto y administra sus imágenes asociadas.

    Permite eliminar imágenes actuales y cargar nuevas respetando límites.
    """
    # Edición administrativa con control de imágenes (máximo y mínimos permitidos).
    producto = get_object_or_404(Product, id=product_id)
    success = False
    errores = {}
    valores = {
        'nombre': producto.nombre,
        'tipo': producto.tipo,
        'tipo_otro': producto.tipo_otro,
        'unidad': producto.unidad,
        'precio': producto.precio,
        'stock': producto.stock,
        'descripcion': producto.descripcion,
        'garantia': producto.garantia,
    }
    existing_images = producto.images.all().order_by('-created_at')
    can_upload_more_images = existing_images.count() < 8
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        tipo = request.POST.get('tipo', '').strip()
        tipo_otro = request.POST.get('tipo_otro', '').strip()
        unidad = request.POST.get('unidad', '').strip()
        precio = request.POST.get('precio', '').strip()
        stock = request.POST.get('stock', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        garantia = request.POST.get('garantia', '').strip()
        # método de pago y entrega eliminados
        delete_image_ids = request.POST.getlist('delete_images')
        new_images = request.FILES.getlist('new_images')
        valores = {
            'nombre': nombre,
            'tipo': tipo,
            'tipo_otro': tipo_otro,
            'unidad': unidad,
            'precio': precio,
            'stock': stock,
            'descripcion': descripcion,
            'garantia': garantia,
        }
        # Validaciones básicas
        if not nombre:
            errores['nombre'] = 'El nombre es obligatorio.'
        if not tipo:
            errores['tipo'] = 'El tipo es obligatorio.'
        if tipo == 'Otros' and not tipo_otro:
            errores['tipo_otro'] = 'Debes especificar el tipo.'
        if not unidad:
            errores['unidad'] = 'La unidad es obligatoria.'
        if not precio:
            errores['precio'] = 'El precio es obligatorio.'
        else:
            try:
                precio_val = Decimal(precio)
                if precio_val <= 0:
                    errores['precio'] = 'El precio debe ser mayor que 0.'
            except (InvalidOperation, ValueError):
                errores['precio'] = 'Precio inválido.'
        if not stock:
            errores['stock'] = 'La cantidad disponible es obligatoria.'
        else:
            try:
                stock_val = int(stock)
                if stock_val < 0:
                    errores['stock'] = 'La cantidad disponible no puede ser negativa.'
            except (TypeError, ValueError):
                errores['stock'] = 'Cantidad disponible inválida.'
        if not descripcion:
            errores['descripcion'] = 'La descripción es obligatoria.'
        if not garantia:
            errores['garantia'] = 'La garantía es obligatoria.'
        # Imágenes
        delete_qs = producto.images.filter(id__in=delete_image_ids)
        existing_count = existing_images.count()
        remaining_count = existing_count - delete_qs.count()
        total_after_update = remaining_count + len(new_images)
        if existing_count >= 8 and len(new_images) > 0 and delete_qs.count() == 0:
            errores['fotos'] = 'Ya tienes 8 imágenes. Elimina alguna para poder subir nuevas.'
        elif total_after_update <= 0:
            errores['fotos'] = 'El producto debe tener al menos una imagen.'
        elif total_after_update > 8:
            errores['fotos'] = 'Solo puedes mantener máximo 8 imágenes por producto.'
        else:
            for photo in new_images:
                if not (photo.content_type or '').startswith('image/'):
                    errores['fotos'] = 'Solo se permiten archivos de imagen.'
                    break
        # Validación de modelo
        if not errores:
            producto.nombre = nombre
            producto.tipo = tipo
            producto.tipo_otro = tipo_otro
            producto.unidad = unidad
            producto.precio = precio
            producto.stock = stock
            producto.descripcion = descripcion
            producto.garantia = garantia
            if int(stock) <= 0:
                producto.is_active = False
            # método de pago y entrega eliminados
            try:
                producto.full_clean()
            except ValidationError as exc:
                for field, messages in exc.message_dict.items():
                    errores[field] = messages[0] if messages else 'Valor inválido.'
            else:
                producto.save()
                if delete_qs.exists():
                    delete_qs.delete()
                for image_file in new_images:
                    ProductImage.objects.create(product=producto, image=image_file)
                success = True
                existing_images = producto.images.all().order_by('-created_at')
                can_upload_more_images = existing_images.count() < 8
    if success:
        return redirect('administrador:producs_page')
    return render(request, 'administrador/producto_editar_admin.html', {
        'success': success,
        'errores': errores,
        'valores': valores,
        'producto': producto,
        'existing_images': existing_images,
        'can_upload_more_images': can_upload_more_images,
        'tipo_choices': Product.TIPO_CHOICES,
        'unidad_choices': Product.UNIDAD_CHOICES,
    })
# Bloquear/desbloquear producto desde admin
from Productos.models import Product
from django.views.decorators.http import require_POST

@require_POST
def producto_admin_block_view(request, product_id):
    """Marca un producto como inactivo."""
    # Cambia el estado del producto a inactivo.
    producto = get_object_or_404(Product, id=product_id)
    producto.is_active = False
    producto.disabled_by_admin = True
    producto.save(update_fields=["is_active", "disabled_by_admin"])
    return JsonResponse({'ok': True})

@require_POST
def producto_admin_unblock_view(request, product_id):
    """Marca un producto como activo."""
    # Cambia el estado del producto a activo.
    producto = get_object_or_404(Product, id=product_id)
    producto.is_active = True
    producto.disabled_by_admin = False
    producto.save(update_fields=["is_active", "disabled_by_admin"])
    return JsonResponse({'ok': True})


def admin_notifications_view(request):
    if not request.user.is_authenticated:
        return redirect('usuarios:login')

    from usuarios.models import Register
    register_admin = Register.objects.filter(id_usuario=request.user.id, estado='admin').first()
    if not register_admin:
        return redirect('usuarios:home_customer')

    from Mensajes.models import AdminNotification

    notifications = AdminNotification.objects.select_related(
        'sender_register',
        'sender_user',
        'product',
    ).order_by('-created_at')

    return render(request, 'administrador/admin_notifications.html', {
        'notifications': notifications,
    })


@require_POST
def admin_notification_mark_read_view(request, notification_id):
    if not request.user.is_authenticated:
        return redirect('usuarios:login')

    from usuarios.models import Register
    register_admin = Register.objects.filter(id_usuario=request.user.id, estado='admin').first()
    if not register_admin:
        return redirect('usuarios:home_customer')

    from Mensajes.models import AdminNotification

    notification = get_object_or_404(AdminNotification, id=notification_id)
    if not notification.is_read:
        notification.is_read = True
        notification.save(update_fields=['is_read'])

    return redirect('administrador:admin_notifications')
# ============================================================
# BLOQUE 4: MENSAJERIA ADMINISTRATIVA
# ============================================================

# Enviar mensaje general (a todos, solo usuarios o solo tiendas)
@require_POST
def usuario_admin_enviar_mensaje_general_view(request):
    """Envía un mensaje masivo a usuarios, tiendas o ambos segmentos.

    El mensaje se registra en AdminToUserMessage y también se intenta enviar por email.
    """
    # Envía mensaje masivo segmentado y registra trazabilidad en base de datos.
    if not request.user.is_authenticated:
        return JsonResponse({'ok': False, 'error': 'No autenticado'}, status=403)

    from usuarios.models import Register

    register_admin = Register.objects.filter(id_usuario=request.user.id, estado='admin').first()
    if not register_admin:
        return JsonResponse({'ok': False, 'error': 'No autorizado'}, status=403)

    import json
    from Mensajes.models import AdminToUserMessage
    from Tiendas.models import Shop
    from django.core.mail import send_mail

    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': 'JSON inválido'}, status=400)

    texto = (data.get('mensaje') or '').strip()
    destinatario = data.get('destinatario', 'all')

    if not texto:
        return JsonResponse({'ok': False, 'error': 'Mensaje vacío'}, status=400)

    usuarios = []
    if destinatario == 'all':
        usuarios = list(Register.objects.all())
    elif destinatario == 'users':
        # Excluir usuarios que son dueños de tienda
        usuarios_con_tienda = Shop.objects.values_list('owner__id', flat=True)
        usuarios = list(Register.objects.exclude(id_usuario__in=usuarios_con_tienda))
    elif destinatario == 'shops':
        # Solo dueños de tienda
        usuarios_con_tienda = Shop.objects.values_list('owner__id', flat=True)
        usuarios = list(Register.objects.filter(id_usuario__in=usuarios_con_tienda))
    else:
        return JsonResponse({'ok': False, 'error': 'Destinatario inválido'}, status=400)

    for usuario in usuarios:
        AdminToUserMessage.objects.create(
            usuario=usuario,
            texto=texto,
            enviado=True
        )
        send_mail(
            'Mensaje importante de Agrophia',
            texto,
            'no-reply@agrophia.com',
            [usuario.correo_electronico],
            fail_silently=True
        )

    return JsonResponse({'ok': True, 'enviados': len(usuarios)})
# ============================================================
# BLOQUE 5: PERFIL ADMIN Y FORMULARIO DE USUARIO ADMIN
# ============================================================

# Vista para editar el perfil del admin autenticado
from django.contrib.auth.decorators import login_required
def admin_editar_perfil_view(request):
    """Permite al admin editar su perfil y datos de tienda relacionados."""
    # Permite al admin actualizar sus datos y, si aplica, datos de su tienda.
    from usuarios.models import Register
    admin_id = request.session.get('admin_user_id')
    if not admin_id:
        return redirect('usuarios:login')
    usuario = get_object_or_404(Register, id_usuario=admin_id)
    tienda = Shop.objects.filter(owner__id=usuario.id_usuario).first()
    if request.method == 'POST':
        form = UsuarioAdminForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            if tienda:
                tienda.nombre = request.POST.get('tienda_nombre', tienda.nombre)
                tienda.email = request.POST.get('tienda_email', tienda.email)
                tienda.telefono = request.POST.get('tienda_telefono', tienda.telefono)
                tienda.departamento = request.POST.get('tienda_departamento', tienda.departamento)
                tienda.municipio = request.POST.get('tienda_municipio', tienda.municipio)
                tienda.direccion = request.POST.get('tienda_direccion', tienda.direccion)
                tienda.descripcion = request.POST.get('tienda_descripcion', tienda.descripcion)
                tienda.horario = request.POST.get('tienda_horario', tienda.horario)
                tienda.punto_fisico = request.POST.get('tienda_punto_fisico', 'True') == 'True'
                tienda.is_active = request.POST.get('tienda_is_active', 'True') == 'True'
                tienda.save()
            return redirect('administrador:home_admin')
    else:
        form = UsuarioAdminForm(instance=usuario)
    return render(request, 'administrador/usuario_editar_admin.html', {'form': form, 'usuario': usuario, 'tienda': tienda, 'es_perfil_admin': True})
from django.shortcuts import render, get_object_or_404, redirect
from django import forms
from django.http import JsonResponse, HttpResponse
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from usuarios.models import Register
from Tiendas.models import Shop
from django.contrib.auth.models import User

class UsuarioAdminForm(forms.ModelForm):
    """Formulario de edición de perfil de usuario en administración."""
    # Formulario técnico para edición de usuarios desde el panel admin.
    contrasena = forms.CharField(label='Contraseña', widget=forms.TextInput(attrs={'type': 'text'}), required=False)

    class Meta:
        from usuarios.models import Register
        model = Register
        fields = [
            'foto', 'tipo_documento', 'numero_documento', 'nombres', 'apellidos',
            'correo_electronico', 'telefono', 'departamento', 'municipio',
            'direccion_completa', 'descripcion_perfil', 'estado', 'contrasena'
        ]
        widgets = {
            'estado': forms.Select(choices=[('activo', 'Activo'), ('inactivo', 'Inactivo')]),
            'descripcion_perfil': forms.Textarea(attrs={'rows': 2}),
        }

    def save(self, commit=True):
        """Persiste el formulario y sincroniza contraseña personalizada si fue enviada."""
        instance = super().save(commit=False)
        if self.cleaned_data.get('contrasena'):
            instance.contrasena = self.cleaned_data['contrasena']
        if commit:
            instance.save()
        return instance

def usuario_admin_editar_view(request, usuario_id):
    """Edita un usuario y, si aplica, su tienda asociada."""
    # Edición integral de usuario y sincronización de tienda asociada.
    from usuarios.models import Register
    usuario = get_object_or_404(Register, id=usuario_id)
    tienda = Shop.objects.filter(owner__id=usuario.id_usuario).first()
    if request.method == 'POST':
        form = UsuarioAdminForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            # Guardar datos de tienda si existen
            if tienda:
                tienda.nombre = request.POST.get('tienda_nombre', tienda.nombre)
                tienda.email = request.POST.get('tienda_email', tienda.email)
                tienda.telefono = request.POST.get('tienda_telefono', tienda.telefono)
                tienda.departamento = request.POST.get('tienda_departamento', tienda.departamento)
                tienda.municipio = request.POST.get('tienda_municipio', tienda.municipio)
                tienda.direccion = request.POST.get('tienda_direccion', tienda.direccion)
                tienda.descripcion = request.POST.get('tienda_descripcion', tienda.descripcion)
                tienda.horario = request.POST.get('tienda_horario', tienda.horario)
                tienda.punto_fisico = request.POST.get('tienda_punto_fisico', 'True') == 'True'
                tienda.is_active = request.POST.get('tienda_is_active', 'True') == 'True'
                tienda.save()
            return redirect('administrador:usuario_admin_detalle', usuario_id=usuario.id)
    else:
        form = UsuarioAdminForm(instance=usuario)
    return render(request, 'administrador/usuario_editar_admin.html', {'form': form, 'usuario': usuario, 'tienda': tienda})
# Vista para detalle de usuario admin
from django.shortcuts import get_object_or_404

def usuario_admin_detalle_view(request, usuario_id):
    """Renderiza detalle de usuario para consulta administrativa."""
    # Vista de solo lectura para perfil de usuario y su tienda vinculada.
    from usuarios.models import Register
    usuario = get_object_or_404(Register, id=usuario_id)
    tienda = Shop.objects.filter(owner__id=usuario.id_usuario).first()
    return render(request, 'administrador/usuario_detalle_admin.html', {'usuario': usuario, 'tienda': tienda})
from django.shortcuts import render, redirect
from django.contrib.auth import login, logout
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse
from datetime import datetime
import csv

def admin_verify_code_view(request):
    pending_user_id = request.session.get('pending_admin_user_id')
    pending_register_id = request.session.get('pending_admin_register_id')
    pending_code = request.session.get('pending_admin_code')
    expires_iso = request.session.get('pending_admin_code_expires_at')

    if not pending_user_id or not pending_register_id or not pending_code or not expires_iso:
        messages.error(request, 'Primero debes iniciar sesion como administrador para generar tu codigo de seguridad.')
        return redirect('usuarios:login')

    try:
        expires_at = datetime.fromisoformat(expires_iso)
    except (TypeError, ValueError):
        request.session.pop('pending_admin_user_id', None)
        request.session.pop('pending_admin_register_id', None)
        request.session.pop('pending_admin_code', None)
        request.session.pop('pending_admin_code_expires_at', None)
        messages.error(request, 'El codigo no es valido. Inicia sesion nuevamente.')
        return redirect('usuarios:login')

    if timezone.now() > expires_at:
        reg_expired = Register.objects.filter(id=pending_register_id).first()
        if reg_expired and reg_expired.admin_code_validated:
            reg_expired.admin_code_validated = False
            reg_expired.save(update_fields=['admin_code_validated'])
        request.session.pop('pending_admin_user_id', None)
        request.session.pop('pending_admin_register_id', None)
        request.session.pop('pending_admin_code', None)
        request.session.pop('pending_admin_code_expires_at', None)
        messages.error(request, 'El codigo expiro. Debes iniciar sesion nuevamente.')
        return redirect('usuarios:login')

    if request.method == 'POST':
        code_input = (request.POST.get('code') or '').strip().upper()
        if len(code_input) != 6:
            messages.error(request, 'El codigo debe tener 6 caracteres.')
            return render(request, 'administrador/admin_verify_code.html', {
                'remaining_seconds': int((expires_at - timezone.now()).total_seconds()),
            })

        if code_input != pending_code:
            messages.error(request, 'Codigo incorrecto.')
            return render(request, 'administrador/admin_verify_code.html', {
                'remaining_seconds': int((expires_at - timezone.now()).total_seconds()),
            })

        user = User.objects.filter(id=pending_user_id).first()
        reg = Register.objects.filter(id=pending_register_id, estado='admin').first()
        if not user or not reg:
            request.session.pop('pending_admin_user_id', None)
            request.session.pop('pending_admin_register_id', None)
            request.session.pop('pending_admin_code', None)
            request.session.pop('pending_admin_code_expires_at', None)
            messages.error(request, 'No se pudo validar el administrador.')
            return redirect('usuarios:login')

        login(request, user, backend='django.contrib.auth.backends.ModelBackend')
        request.session['admin_user_id'] = user.id
        reg.admin_code_validated = True
        reg.save(update_fields=['admin_code_validated'])

        request.session.pop('pending_admin_user_id', None)
        request.session.pop('pending_admin_register_id', None)
        request.session.pop('pending_admin_code', None)
        request.session.pop('pending_admin_code_expires_at', None)
        next_admin_path = request.session.pop('pending_admin_next', '/administrador/home/')
        if not next_admin_path.startswith('/administrador/'):
            next_admin_path = '/administrador/home/'

        return redirect(next_admin_path)

    return render(request, 'administrador/admin_verify_code.html', {
        'remaining_seconds': max(0, int((expires_at - timezone.now()).total_seconds())),
    })


def admin_logout_view(request):
    """Cierra sesión admin y limpia banderas de validación del código admin."""
    # Logout limpio + limpieza de bandera de validación de código admin.
    # Al cerrar sesión, limpiar validación en la base de datos
    user_id = request.session.get('admin_user_id')
    if user_id:
        try:
            reg = Register.objects.get(id_usuario=user_id)
            reg.admin_code_validated = False
            reg.save()
        except Register.DoesNotExist:
            pass
    logout(request)
    request.session.flush()
    return redirect('/')

# ============================================================
# BLOQUE 7: UTILIDADES DE ACTIVIDAD Y ORDENAMIENTO
# ============================================================

def safe_fecha(fecha):
    """Devuelve una fecha formateada de manera segura para UI/reportes."""
    # Formatea fecha de forma segura para UI/reportes.
    if hasattr(fecha, 'strftime'):
        return fecha.strftime('%d/%m/%Y %H:%M')
    return 'Sin fecha'

def activity_category_label(category):
    """Mapea la categoría interna de actividad a una etiqueta visible."""
    # Convierte clave técnica de categoría a etiqueta legible.
    labels = {
        'all': 'Toda la actividad',
        'users': 'Usuarios',
        'shops': 'Tiendas',
        'products': 'Productos',
        'orders': 'Pedidos',
    }
    return labels.get(category, 'Actividad')

def _activity_timestamp(fecha):
    """Convierte una fecha a timestamp numérico para ordenamiento estable."""
    # Obtiene timestamp numérico para llaves de ordenamiento robustas.
    if hasattr(fecha, 'timestamp'):
        return fecha.timestamp()
    return 0

def activity_latest_sort_key(evento):
    """Clave de ordenamiento para eventos del más reciente al más antiguo."""
    # Orden descendente por fecha (más reciente primero).
    fecha = evento.get('occurred_at')
    return (0 if fecha else 1, -_activity_timestamp(fecha), evento.get('sequence', 0))

def activity_first_sort_key(evento):
    """Clave de ordenamiento para eventos del más antiguo al más reciente."""
    # Orden ascendente por fecha (más antiguo primero).
    fecha = evento.get('occurred_at')
    return (0 if fecha else 1, _activity_timestamp(fecha), evento.get('sequence', 0))

def build_activity_events():
    """Construye una colección unificada de eventos del sistema.

    Fuentes incluidas:
    - Register (altas y bloqueos).
    - Shop (creación y deshabilitación).
    - Product (creación y deshabilitación).
    - Order (nuevos pedidos).
    """
    # Construye una línea de tiempo unificada desde usuarios, tiendas, productos y pedidos.
    eventos = []
    sequence = 0

    registros = list(Register.objects.exclude(estado='admin').order_by('-id'))
    usuarios_auth = User.objects.in_bulk([
        registro.id_usuario for registro in registros if registro.id_usuario
    ])

    for registro in registros:
        usuario_auth = usuarios_auth.get(registro.id_usuario)
        fecha_registro = getattr(usuario_auth, 'date_joined', None)
        eventos.append({
            'sequence': sequence,
            'occurred_at': fecha_registro,
            'category': 'users',
            'tipo': 'Registro de usuario',
            'descripcion': f"Nuevo usuario registrado: {registro.nombres} {registro.apellidos} ({registro.numero_documento})",
        })
        sequence += 1

        if registro.estado == 'inactivo':
            eventos.append({
                'sequence': sequence,
                'occurred_at': None,
                'category': 'users',
                'tipo': 'Usuario bloqueado',
                'descripcion': f"Usuario bloqueado: {registro.nombres} {registro.apellidos} ({registro.numero_documento})",
            })
            sequence += 1

    for tienda in Shop.objects.select_related('owner').order_by('-created_at'):
        propietario = tienda.owner.username if tienda.owner else 'Sin propietario'
        eventos.append({
            'sequence': sequence,
            'occurred_at': tienda.created_at,
            'category': 'shops',
            'tipo': 'Tienda creada',
            'descripcion': f"Nueva tienda creada: {tienda.nombre} (Dueño: {propietario})",
        })
        sequence += 1

        if not tienda.is_active:
            eventos.append({
                'sequence': sequence,
                'occurred_at': None,
                'category': 'shops',
                'tipo': 'Tienda deshabilitada',
                'descripcion': f"Tienda deshabilitada: {tienda.nombre} (Dueño: {propietario})",
            })
            sequence += 1

    for producto in Product.objects.select_related('owner', 'shop').order_by('-created_at'):
        propietario = producto.owner.username if producto.owner else 'Sin propietario'
        eventos.append({
            'sequence': sequence,
            'occurred_at': producto.created_at,
            'category': 'products',
            'tipo': 'Producto creado',
            'descripcion': f"Producto creado: {producto.nombre} (Productor: {propietario})",
        })
        sequence += 1

        if not producto.is_active:
            eventos.append({
                'sequence': sequence,
                'occurred_at': None,
                'category': 'products',
                'tipo': 'Producto deshabilitado',
                'descripcion': f"Producto deshabilitado: {producto.nombre} (ID: {producto.id})",
            })
            sequence += 1

    for pedido in Order.objects.select_related('customer').order_by('-created_at'):
        eventos.append({
            'sequence': sequence,
            'occurred_at': pedido.created_at,
            'category': 'orders',
            'tipo': 'Pedido realizado',
            'descripcion': f"Nuevo pedido #{pedido.id} realizado por {pedido.customer.username} por ${pedido.total_amount}",
        })
        sequence += 1

    return eventos

def filter_activity_events(events, category='all', scope='all', count_mode='latest', count_value=100, period='month'):
    """Filtra eventos según criterios de reporte.

    Args:
        events: Lista de eventos preconstruidos.
        category: Categoria deseada (all/users/shops/products/orders).
        scope: Alcance (all/count/period).
        count_mode: latest o first.
        count_value: Cantidad a retornar cuando scope=count.
        period: month o year para scope=period.
    """
    # Filtra y recorta eventos según categoría, alcance temporal y modo de conteo.
    valid_categories = {'all', 'users', 'shops', 'products', 'orders'}
    valid_scopes = {'all', 'count', 'period'}
    valid_count_modes = {'latest', 'first'}
    valid_periods = {'month', 'year'}

    if category not in valid_categories:
        category = 'all'
    if scope not in valid_scopes:
        scope = 'all'
    if count_mode not in valid_count_modes:
        count_mode = 'latest'
    if period not in valid_periods:
        period = 'month'

    filtrados = [
        evento for evento in events
        if category == 'all' or evento['category'] == category
    ]

    if scope == 'period':
        ahora = timezone.now()
        if period == 'month':
            filtrados = [
                evento for evento in filtrados
                if evento['occurred_at']
                and evento['occurred_at'].year == ahora.year
                and evento['occurred_at'].month == ahora.month
            ]
        else:
            filtrados = [
                evento for evento in filtrados
                if evento['occurred_at'] and evento['occurred_at'].year == ahora.year
            ]
        return sorted(filtrados, key=activity_latest_sort_key)

    if scope == 'count':
        try:
            count = int(count_value)
        except (TypeError, ValueError):
            count = 100
        count = max(1, min(count, 1000))
        sort_key = activity_latest_sort_key if count_mode == 'latest' else activity_first_sort_key
        return sorted(filtrados, key=sort_key)[:count]

    return sorted(filtrados, key=activity_latest_sort_key)

def activity_scope_label(scope, count_mode='latest', count_value=100, period='month'):
    """Genera el subtítulo legible del alcance aplicado al reporte."""
    # Etiqueta human-readable del alcance aplicado al reporte de actividad.
    if scope == 'count':
        prefix = 'Ultimas' if count_mode == 'latest' else 'Primeras'
        return f"{prefix} {count_value} actividades"
    if scope == 'period':
        return 'Actividades de este mes' if period == 'month' else 'Actividades de este año'
    return 'Toda la actividad disponible'

def autofit_worksheet_columns(ws):
    """Ajusta anchos de columna en una hoja openpyxl según contenido."""
    # Ajusta automáticamente anchos de columna para exportaciones Excel.
    from openpyxl.utils import get_column_letter

    for col_index, col in enumerate(ws.iter_cols(min_col=1, max_col=ws.max_column), start=1):
        max_length = 0
        for cell in col:
            try:
                cell_value = '' if cell.value is None else str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_index)].width = min(max_length + 2, 60)

def render_generic_report_print(request, title, headers, rows, subtitle=''):
    """Renderiza una plantilla de impresión genérica para reportes tabulares."""
    # Render genérico para reportes imprimibles en HTML.
    return render(request, 'administrador/reporte_generico_print.html', {
        'title': title,
        'headers': headers,
        'rows': rows,
        'subtitle': subtitle,
    })

# ============================================================
# BLOQUE 8: HOME ADMIN Y REPORTE DE ACTIVIDAD RECIENTE
# ============================================================

def home_admin_view(request):
    """Renderiza el home admin con los últimos eventos de actividad."""
    # Home administrativa mostrando resumen de actividad reciente.
    if not request.user.is_authenticated:
        return redirect('usuarios:login')

    register_admin = Register.objects.filter(id_usuario=request.user.id, estado='admin').first()
    if not register_admin:
        logout(request)
        request.session.flush()
        return redirect('usuarios:login')

    if request.session.get('admin_user_id') != request.user.id or not register_admin.admin_code_validated:
        messages.error(request, 'Debes completar la verificacion de seguridad para ingresar al panel.')
        return redirect('administrador:admin_verify_code')

    eventos = [
        {
            'fecha': safe_fecha(evento['occurred_at']),
            'tipo': evento['tipo'],
            'descripcion': evento['descripcion'],
        }
        for evento in filter_activity_events(
            build_activity_events(),
            scope='count',
            count_mode='latest',
            count_value=10,
        )
    ]
    return render(request, 'administrador/home_admin.html', {'eventos': eventos})

def reporte_actividad_reciente_view(request):
    """Genera reporte de actividad reciente en impresión o Excel.

    Admite filtros por categoría, alcance, periodo y modo de conteo.
    """
    # Genera reporte de actividad reciente (Excel o impresión).
    activity_type = request.GET.get('activity_type', 'all')
    scope = request.GET.get('scope', 'all')
    count_mode = request.GET.get('count_mode', 'latest')
    count_value = request.GET.get('count_value', '100')
    period = request.GET.get('period', 'month')
    output_format = request.GET.get('output_format', 'excel')

    eventos = filter_activity_events(
        build_activity_events(),
        category=activity_type,
        scope=scope,
        count_mode=count_mode,
        count_value=count_value,
        period=period,
    )

    resumen = {
        'tipo_actividad': activity_category_label(activity_type),
        'alcance': activity_scope_label(scope, count_mode, count_value, period),
        'total_resultados': len(eventos),
    }

    if output_format == 'print':
        return render(request, 'administrador/reporte_actividad_reciente_print.html', {
            'eventos': [
                {
                    'fecha': safe_fecha(evento['occurred_at']),
                    'categoria': activity_category_label(evento['category']),
                    'tipo': evento['tipo'],
                    'descripcion': evento['descripcion'],
                }
                for evento in eventos
            ],
            'resumen': resumen,
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Actividad reciente'

    ws.merge_cells('A1:D1')
    ws['A1'] = 'Reporte de actividad reciente'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='16A34A', end_color='16A34A', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.append([])
    ws.append(['Tipo de actividad', resumen['tipo_actividad'], 'Alcance', resumen['alcance']])
    ws.append(['Resultados', resumen['total_resultados'], '', ''])
    ws.append([])

    headers = ['Fecha', 'Categoria', 'Tipo', 'Descripcion']
    ws.append(headers)
    header_row_index = ws.max_row
    for cell in ws[header_row_index]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for evento in eventos:
        ws.append([
            safe_fecha(evento['occurred_at']),
            activity_category_label(evento['category']),
            evento['tipo'],
            evento['descripcion'],
        ])
    autofit_worksheet_columns(ws)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="actividad_reciente_agrophia.xlsx"'
    wb.save(response)
    return response

# ============================================================
# BLOQUE 9: USUARIOS, PEDIDOS Y REPORTES ADMIN
# ============================================================

def usuarios_admin_view(request):
    """Renderiza listado de usuarios administrables (excluye admins)."""
    # Lista general de usuarios no administradores.
    usuarios = Register.objects.exclude(estado='admin').order_by('nombres', 'apellidos')
    return render(request, 'administrador/usuarios_admin.html', {'usuarios': usuarios})

def orders_page_view(request):
    """Renderiza vista administrativa de pedidos."""
    # Vista tabla de pedidos para administración.
    pedidos = Order.objects.select_related('customer').order_by('-created_at')
    return render(request, 'administrador/orders_page.html', {'pedidos': pedidos})

def producs_page_view(request):
    """Renderiza vista administrativa de productos."""
    # Vista tabla de productos para administración.
    productos = Product.objects.all().order_by('-created_at')
    return render(request, 'administrador/producs_page.html', {'productos': productos})

def usuario_admin_block_view(request, usuario_id):
    """Bloquea un usuario (estado=inactivo) vía endpoint POST."""
    # Bloquea usuario cambiando estado a inactivo.
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=400)

    usuario = get_object_or_404(Register, id=usuario_id)
    usuario.estado = 'inactivo'
    usuario.save(update_fields=['estado'])
    return JsonResponse({'ok': True})

def usuario_admin_unblock_view(request, usuario_id):
    """Desbloquea un usuario (estado=activo) vía endpoint POST."""
    # Desbloquea usuario cambiando estado a activo.
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=400)

    usuario = get_object_or_404(Register, id=usuario_id)
    usuario.estado = 'activo'
    usuario.save(update_fields=['estado'])
    return JsonResponse({'ok': True})

def usuario_admin_enviar_mensaje_view(request, usuario_id):
    """Envía mensaje administrativo individual a un usuario."""
    # Envío individual de mensaje administrativo a un usuario.
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=400)

    import json
    from Mensajes.models import AdminToUserMessage
    from django.core.mail import send_mail

    usuario = get_object_or_404(Register, id=usuario_id)

    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        data = {}

    texto = (data.get('mensaje') or '').strip()
    if not texto:
        return JsonResponse({'ok': False, 'error': 'Mensaje vacío'}, status=400)

    AdminToUserMessage.objects.create(
        usuario=usuario,
        texto=texto,
        enviado=True,
    )
    send_mail(
        'Mensaje importante de Agrophia',
        texto,
        'no-reply@agrophia.com',
        [usuario.correo_electronico],
        fail_silently=True,
    )
    return JsonResponse({'ok': True})

def reporte_usuarios_view(request):
    """Genera reporte de usuarios en Excel o impresión."""
    # Exporta listado de usuarios en Excel o versión imprimible.
    output_format = request.GET.get('output_format', 'excel')
    usuarios = Register.objects.exclude(estado='admin').order_by('nombres', 'apellidos')
    headers = ['ID', 'Nombres', 'Apellidos', 'Correo', 'Teléfono', 'Departamento', 'Municipio', 'Estado']
    rows = [
        [u.id, u.nombres, u.apellidos, u.correo_electronico, u.telefono, u.departamento, u.municipio, u.estado]
        for u in usuarios
    ]

    if output_format == 'print':
        return render_generic_report_print(request, 'Reporte de usuarios', headers, rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Usuarios Agrophia'
    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='16A34A', end_color='16A34A', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        ws.append(row)
    autofit_worksheet_columns(ws)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="usuarios_agrophia.xlsx"'
    wb.save(response)
    return response

def usuario_admin_crear_view(request):
    """Crea un usuario desde admin con validaciones de datos y duplicados."""
    # Registro manual de usuario desde panel administrativo.
    from usuarios.models import Register
    from django.contrib.auth.models import User
    success = False
    errores = {}
    valores = {}
    if request.method == 'POST':
        from Administrador.departamentos_municipios import DEPARTAMENTOS_MUNICIPIOS
        nombres = request.POST.get('nombres', '').strip()
        apellidos = request.POST.get('apellidos', '').strip()
        tipo_documento = request.POST.get('tipo_documento', '').strip()
        numero_documento = request.POST.get('numero_documento', '').strip()
        correo_electronico = request.POST.get('correo_electronico', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        departamento = request.POST.get('departamento', '').strip()
        municipio = request.POST.get('municipio', '').strip()
        direccion_completa = request.POST.get('direccion_completa', '').strip()
        contrasena = request.POST.get('contrasena', '').strip()
        valores = {
            'nombres': nombres,
            'apellidos': apellidos,
            'tipo_documento': tipo_documento,
            'numero_documento': numero_documento,
            'correo_electronico': correo_electronico,
            'telefono': telefono,
            'departamento': departamento,
            'municipio': municipio,
            'direccion_completa': direccion_completa,
        }
        # Validaciones
        if not nombres:
            errores['nombres'] = 'El nombre es obligatorio.'
        if not apellidos:
            errores['apellidos'] = 'El apellido es obligatorio.'
        if not tipo_documento:
            errores['tipo_documento'] = 'El tipo de documento es obligatorio.'
        if not numero_documento:
            errores['numero_documento'] = 'El número de documento es obligatorio.'
        elif Register.objects.filter(numero_documento=numero_documento).exists():
            errores['numero_documento'] = 'Ya existe un usuario con ese número de documento.'
        if not correo_electronico:
            errores['correo_electronico'] = 'El correo es obligatorio.'
        elif Register.objects.filter(correo_electronico=correo_electronico).exists():
            errores['correo_electronico'] = 'Ya existe un usuario con ese correo.'
        if not telefono:
            errores['telefono'] = 'El teléfono es obligatorio.'
        if not departamento:
            errores['departamento'] = 'El departamento es obligatorio.'
        elif departamento not in DEPARTAMENTOS_MUNICIPIOS:
            errores['departamento'] = 'Departamento no válido.'
        if not municipio:
            errores['municipio'] = 'El municipio es obligatorio.'
        elif departamento in DEPARTAMENTOS_MUNICIPIOS and municipio not in DEPARTAMENTOS_MUNICIPIOS[departamento]:
            errores['municipio'] = 'El municipio no corresponde al departamento seleccionado.'
        if not direccion_completa:
            errores['direccion_completa'] = 'La dirección es obligatoria.'
        if not contrasena or len(contrasena) < 8:
            errores['contrasena'] = 'La contraseña es obligatoria y debe tener al menos 8 caracteres.'
        if not errores:
            user, created = User.objects.get_or_create(
                username=numero_documento,
                defaults={
                    'email': correo_electronico,
                    'first_name': nombres,
                    'last_name': apellidos
                }
            )
            Register.objects.create(
                id_usuario=user.id,
                nombres=nombres,
                apellidos=apellidos,
                tipo_documento=tipo_documento,
                numero_documento=numero_documento,
                correo_electronico=correo_electronico,
                telefono=telefono,
                departamento=departamento,
                municipio=municipio,
                direccion_completa=direccion_completa,
                contrasena=contrasena,
                estado='activo'
            )
            success = True
            valores = {}
    from Administrador.departamentos_municipios import DEPARTAMENTOS_MUNICIPIOS
    return render(request, 'administrador/usuario_crear_admin.html', {
        'success': success,
        'errores': errores,
        'valores': valores,
        'departamentos_municipios': DEPARTAMENTOS_MUNICIPIOS
    })

from Pedidos.models import Order, OrderItem
from django.shortcuts import get_object_or_404, render

# ============================================================
# BLOQUE 10: PEDIDOS (DETALLE / EDICION / REPORTE)
# ============================================================

def pedido_admin_detalle_view(request, pedido_id):
    """Muestra detalle de un pedido y sus ítems asociados."""
    # Detalle de pedido y sus ítems con producto/agricultor relacionados.
    pedido = get_object_or_404(Order, id=pedido_id)
    items = pedido.items.select_related('product', 'farmer').all()
    return render(request, 'administrador/pedido_detalle_card.html', {
        'pedido': pedido,
        'items': items,
    })

def pedido_admin_editar_view(request, pedido_id):
    """Permite editar estado y dirección de entrega de un pedido."""
    # Edición administrativa del estado y dirección de entrega del pedido.
    pedido = get_object_or_404(Order, id=pedido_id)
    errores = {}
    if request.method == 'POST':
        status = request.POST.get('status', '').strip()
        delivery_address = request.POST.get('delivery_address', '').strip()
        # Validaciones básicas
        if status not in dict(Order.STATUS_CHOICES):
            errores['status'] = 'Estado inválido.'
        if not delivery_address:
            errores['delivery_address'] = 'La dirección de entrega es obligatoria.'
        if not errores:
            pedido.status = status
            pedido.delivery_address = delivery_address
            pedido.save()
            return redirect('administrador:orders_page')
    items = pedido.items.select_related('product', 'farmer').all()
    return render(request, 'administrador/pedido_editar_card.html', {
        'pedido': pedido,
        'items': items,
        'errores': errores,
        'status_choices': Order.STATUS_CHOICES,
    })

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

def reporte_pedidos_view(request):
    """Genera reporte de pedidos filtrable por estado (Excel o impresión)."""
    # Reporte de pedidos filtrable por estado (Excel/print).
    estado = request.GET.get('estado', 'todos')
    output_format = request.GET.get('output_format', 'excel')
    queryset = Order.objects.select_related('customer').all()
    if estado and estado != 'todos':
        queryset = queryset.filter(status=estado)
    headers = ['ID', 'Cliente', 'Fecha', 'Total', 'Estado', 'Dirección de entrega']
    rows = [
        [
            p.id,
            p.customer.get_full_name() if hasattr(p.customer, 'get_full_name') else p.customer.username,
            p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else '',
            str(p.total_amount),
            p.get_status_display(),
            p.delivery_address or '',
        ]
        for p in queryset.order_by('-created_at')
    ]

    if output_format == 'print':
        subtitle = 'Estado: Todos' if estado == 'todos' else f'Estado: {estado}'
        return render_generic_report_print(request, 'Reporte de pedidos', headers, rows, subtitle=subtitle)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pedidos Agrophia'
    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='16A34A', end_color='16A34A', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        ws.append(row)
    autofit_worksheet_columns(ws)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pedidos_agrophia.xlsx"'
    wb.save(response)
    return response
